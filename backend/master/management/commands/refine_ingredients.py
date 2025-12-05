import openpyxl
import os
from django.core.management.base import BaseCommand
from master.models import IngredientMaster
from django.db.models import Q

class Command(BaseCommand):
    help = 'Refine existing ingredients and import from Excel with better categorization'

    def handle(self, *args, **options):
        self.stdout.write('🔄 Starting ingredient refinement...')

        # 1. 기존 데이터 재분류
        self.refine_existing_data()

        # 2. 엑셀 파일 임포트
        self.import_excel_data()

    def get_category_and_icon(self, name):
        """이름을 분석하여 카테고리와 아이콘을 반환"""
        name = name.replace(' ', '') # 공백 제거 후 분석
        
        # 1. 육류
        if any(x in name for x in ['돼지', '소고기', '닭', '오리', '양고기', '삼겹살', '목살', '갈비', '한우', '육회']):
            return '육류', '🥩'
        
        # 2. 수산물
        if any(x in name for x in ['생선', '고등어', '갈치', '오징어', '새우', '조개', '멸치', '김', '미역', '다시마', '참치', '연어', '광어', '게', '랍스터']):
            return '수산물', '🐟'
        
        # 3. 채소
        if any(x in name for x in ['배추', '무', '상추', '깻잎', '시금치', '파', '양파', '마늘', '고추', '당근', '오이', '호박', '가지', '버섯', '콩나물', '숙주', '브로콜리', '양배추', '감자', '고구마']):
            return '채소', '🥬'
        
        # 4. 과일
        if any(x in name for x in ['사과', '배', '포도', '딸기', '바나나', '귤', '오렌지', '수박', '참외', '복숭아', '자두', '레몬', '토마토', '키위', '망고']):
            return '과일', '🍎'
        
        # 5. 유제품
        if any(x in name for x in ['우유', '치즈', '요거트', '버터', '크림', '분유']):
            return '유제품', '🥛'
        
        # 6. 곡류/견과류
        if any(x in name for x in ['쌀', '밥', '현미', '찹쌀', '보리', '콩', '팥', '밀가루', '두부', '견과', '아몬드', '호두']):
            return '곡류', '🍚'
            
        # 7. 가공식품
        if any(x in name for x in ['햄', '소시지', '베이컨', '만두', '라면', '통조림', '과자', '빵', '케이크', '소스', '장', '기름', '식용유', '참기름', '마요네즈', '케첩', '초콜릿', '사탕', '젤리', '음료', '주스']):
            return '가공식품', '🥫'

        # 기본값 (기존 분류가 '원재료성식품'이면 '기타'로 변경, 아니면 유지)
        return None, '🥘'

    def refine_existing_data(self):
        self.stdout.write('🛠️ Re-categorizing existing data...')
        ingredients = IngredientMaster.objects.all()
        updated_count = 0

        for ing in ingredients:
            new_category, new_icon = self.get_category_and_icon(ing.name)
            
            needs_save = False
            
            # 카테고리 업데이트 (새 분류가 감지되었거나, 기존 분류가 너무 모호할 때)
            if new_category:
                if ing.category != new_category:
                    ing.category = new_category
                    needs_save = True
            elif ing.category in ['원재료성식품', '농축산물']:
                # 분류를 못 찾았는데 기존 분류가 모호하면 '기타'로 변경
                ing.category = '기타'
                needs_save = True

            # 아이콘 업데이트
            if new_icon and ing.icon != new_icon:
                ing.icon = new_icon
                needs_save = True

            if needs_save:
                ing.save()
                updated_count += 1
        
        self.stdout.write(f'✅ Updated {updated_count} existing ingredients.')

    def import_excel_data(self):
        file_path = 'data/foodDB.xlsx'
        if not os.path.exists(file_path):
            self.stdout.write(self.style.WARNING(f'File not found: {file_path}. Skipping import.'))
            return

        self.stdout.write('📥 Importing from foodDB.xlsx...')
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet = wb.active
            
            # 헤더 확인 (첫 번째 행)
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            
            # 컬럼 인덱스 찾기
            try:
                name_idx = headers.index('식품명')
                cat_idx = headers.index('식품대분류명')
            except ValueError:
                # 컬럼명이 다를 경우 인덱스로 추정 (보통 0, 1, 2...)
                # 사용자가 제공한 정보가 없으므로 일단 5번째(식품명), 3번째(분류) 등으로 추측하거나
                # 앞서 실행한 명령의 결과를 보고 수정해야 함.
                # 일단 안전하게 이름이 포함된 컬럼을 찾음
                name_idx = -1
                cat_idx = -1
                for i, h in enumerate(headers):
                    if h and '식품명' in str(h): name_idx = i
                    if h and '분류' in str(h): cat_idx = i
            
            if name_idx == -1:
                self.stdout.write(self.style.ERROR('Could not find "식품명" column.'))
                return

            created_count = 0
            seen_names = set(IngredientMaster.objects.values_list('name', flat=True))

            for row in sheet.iter_rows(min_row=2, values_only=True):
                raw_name = row[name_idx]
                if not raw_name: continue
                
                # 이름 정제
                clean_name = str(raw_name).split('_')[0].split('(')[0].strip()
                
                if len(clean_name) < 1 or clean_name in seen_names:
                    continue

                # 카테고리 결정
                category, icon = self.get_category_and_icon(clean_name)
                
                if not category:
                    # 엑셀의 분류 사용
                    raw_cat = row[cat_idx] if cat_idx != -1 else '기타'
                    if raw_cat in ['원재료성식품', '농축산물']:
                        category = '기타'
                    else:
                        category = raw_cat or '기타'

                IngredientMaster.objects.create(
                    name=clean_name,
                    category=category,
                    default_unit='개',
                    icon=icon,
                    api_source='foodDB_excel'
                )
                
                seen_names.add(clean_name)
                created_count += 1
                
                if created_count % 100 == 0:
                    self.stdout.write(f'Imported {created_count} items...')

            self.stdout.write(self.style.SUCCESS(f'✅ Successfully imported {created_count} new items from Excel!'))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error reading Excel: {str(e)}'))
